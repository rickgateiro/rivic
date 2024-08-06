import pandas as pd
from tkinter import Tk, filedialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.cell import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from flask import Flask, render_template, request, send_file


# Check if the user selected a file or canceled the selection
if not file_path:
    print("No file selected. Exiting the program.")
else:
    # Read the CSV file
    df = pd.read_csv(file_path)

    # Drop the "AÇÃO" column (if it exists)
    if "AÇÃO" in df.columns:
        df.drop(columns=["AÇÃO"], inplace=True)

    # Create a new DataFrame for the Excel output
    output_df = df.copy()

    # Create the link column by combining the base link and the ID from the "ID" column
    base_link = "https://app08.virtuaserver.com.br/astrum/public/site/html/tipologiaCadastra.html?&"
    output_df["Link"] = base_link + df["id"].astype(str) + "#"

    # Create a new Excel file using openpyxl
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Document List"

    # Add titles to the first row
    titles = ["STATUS", "MACROPROCESSO", "DOCUMENTO", "ID", "LINK"]
    worksheet.append(titles)

    # Write the data from the DataFrame to the Excel file
    for row in dataframe_to_rows(output_df, index=False, header=False):
        worksheet.append(row)

    # Set the header row in bold
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    # Adjust the column width to fit the content
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    # Insert a blank row after the first row
    worksheet.insert_rows(2)

    # Make the "LINK" column clickable
    for row in worksheet.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.hyperlink = cell.value
            cell.font = Font(underline="single", color="0563C1")

    # Save the Excel file
    workbook.save("document_list.xlsx")  # Replace "document_list.xlsx" with the desired output file name

    print("Excel file generated successfully.")
