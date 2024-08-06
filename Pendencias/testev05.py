from flask import Flask, render_template, request, send_file
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# Criar a janela de diálogo para escolher o arquivo CSV
root = tk.Tk()
root.withdraw()  # Esconder a janela principal

# Abrir a caixa de diálogo para escolher o arquivo CSV
file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])

# Verificar se o usuário escolheu um arquivo ou cancelou a seleção
if not file_path:
    print("Nenhum arquivo selecionado. Encerrando o programa.")
else:
    # Ler o arquivo CSV com o pandas
    dataframe = pd.read_csv(file_path)

# Preencher valores vazios na coluna 'Nome operador (avaliação)'
dataframe['Nome operador (identificação)'].fillna('Verificar qual operador identificou', inplace=True)
dataframe['Nome operador (avaliação)'].fillna('Sem operador analisando ainda', inplace=True)

# Identificar valores únicos na coluna "Macroprocesso"
macroprocessos = dataframe['Macroprocesso'].unique()

# Criar um arquivo Excel
workbook = Workbook()

# Remover a planilha padrão
default_sheet = workbook['Sheet']
workbook.remove(default_sheet)

for macroprocesso in macroprocessos:
    # Filtrar o DataFrame para o Macroprocesso atual
    df_macroprocesso = dataframe[dataframe['Macroprocesso'] == macroprocesso]

    # Limitar o nome da aba a 15 caracteres
    aba_nome = macroprocesso[:15]

    # Agrupar os dados por situação e operador de identificação
    agrupado = df_macroprocesso.groupby(['Situação', 'Nome operador (identificação)', 'Nome operador (avaliação)']).size().reset_index(name='Qtd. de Docs')

    # Adicionar coluna 'Status' com valor vazio
    agrupado['Status'] = ''

    # Adicionar coluna 'Macroprocesso' com valor único
    agrupado['Macroprocesso'] = macroprocesso

    # Mostrar os usuários atribuídos a cada situação
    for index, row in agrupado.iterrows():
        macroprocesso = row['Macroprocesso']
        operadorid = row['Nome operador (identificação)']
        contagem = row['Qtd. de Docs']
        situacao = row['Situação']
        status = row['Status']
        print(f"Contagem: {contagem}, Situação: {situacao}")

    # Selecionar apenas as colunas desejadas
    colunas_desejadas = ['Macroprocesso', 'Situação', 'Nome operador (identificação)', 'Nome operador (avaliação)',
                         'Qtd. de Docs', 'Status']
    agrupado = agrupado[colunas_desejadas]

    # Verificar a situação 'ANÁLISE DE AVALIAÇÃO' e criar a coluna 'STATUS' com a mensagem adequada para cada status
    # agrupado['Status'] = ''

    agrupado.loc[agrupado['Situação'] == 'ANÁLISE DE AVALIAÇÃO', 'Status'] = 'ANÁLISE RIVIC'
    agrupado.loc[agrupado['Situação'] == 'EM AVALIAÇÃO', 'Status'] = 'PENDENTE'
    agrupado.loc[agrupado['Situação'] == 'APROVADA', 'Status'] = 'PENDENTE'
    agrupado.loc[agrupado['Situação'] == 'APROVAÇÃO DE AVALIAÇÃO', 'Status'] = 'FINALIZADA'
    agrupado.loc[agrupado['Situação'] == 'FINALIZADA', 'Status'] = 'FINALIZADA'

    # Criar uma aba com o nome limitado do Macroprocesso
    worksheet = workbook.create_sheet(title=aba_nome)

    # Preencher o arquivo Excel com os dados do DataFrame
    for row in agrupado.itertuples(index=False):
        worksheet.append(row)
    for row in worksheet.iter_rows(min_row=2, min_col=colunas_desejadas.index('Status') + 1,
                                   max_col=colunas_desejadas.index('Status') + 1):
        for cell in row:
            if cell.value == 'FINALIZADA':
                cell.fill = PatternFill(start_color='9AFF9A', end_color='9AFF9A', fill_type='solid')
            if cell.value == 'ANÁLISE RIVIC':
                cell.fill = PatternFill(start_color='FF7F00', end_color='FF7F00', fill_type='solid')
            if cell.value == 'PENDENTE':
                cell.fill = PatternFill(start_color='FF3030', end_color='FF3030', fill_type='solid')

    # Adicionar título para cada coluna no topo da planilha
    for col_num, col_name in enumerate(colunas_desejadas, 1):
        worksheet.cell(row=1, column=col_num, value=col_name)
        worksheet.cell(row=1, column=col_num).alignment = Alignment(horizontal='center', vertical='center')

        # Adicionar título em negrito para cada coluna no topo da planilha
        for col_num, col_name in enumerate(colunas_desejadas, 1):
            worksheet.cell(row=1, column=col_num, value=col_name)
            worksheet.cell(row=1, column=col_num).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=1, column=col_num).font = Font(bold=True)

    # Definir a referência da área com os filtros (todas as colunas, primeira linha até última linha)
    ref = f"A1:{worksheet.cell(row=worksheet.max_row, column=len(colunas_desejadas)).coordinate}"
    worksheet.auto_filter.ref = ref

    # Ajustar a largura das colunas selecionadas com base no conteúdo
    for col in worksheet.iter_cols(min_col=1, max_col=len(colunas_desejadas)):
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Centralizar o conteúdo das células
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    primeira_celula = df_macroprocesso['Macroprocesso'].iloc[0]

    # Definir o estilo de borda
    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    # Percorrer todas as células da planilha e aplicar o estilo de borda
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border

# Salvar o arquivo Excel após o loop
workbook.save('Relatorio_Macroprocessos.xlsx')

# Criar a janela para a mensagem de sucesso
janela_sucesso = tk.Tk()
janela_sucesso.withdraw()  # Esconder a janela principal

# Mostrar a caixa de diálogo com a mensagem de sucesso
messagebox.showinfo("Sucesso", "Planilha gerada com sucesso!")

# Fechar a janela de sucesso e encerrar o programa
janela_sucesso.destroy()

