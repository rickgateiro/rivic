from flask import Flask, render_template, request, send_file
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

#VERSÃO 0.6  - Agora vários macroprocessos podem ser gerados juntos, criando-se uma aba para cada um.
#dia 12/09 - Corrigio o bug na linha suprimida


app = Flask(__name__)

def criar_aba(workbook, nome_aba, abas_criadas):
    aba_nome_base = extrair_identificador(nome_aba)

    if aba_nome_base not in abas_criadas:
        abas_criadas.add(aba_nome_base)
        return workbook.create_sheet(title=aba_nome_base)
    else:
        # Se a aba já existir, gera um novo nome exclusivo
        novo_nome = criar_nome_aba(workbook, aba_nome_base)
        return workbook.create_sheet(title=novo_nome)

def extrair_identificador(nome_aba):
    # Adapte isso conforme a lógica específica para a sua extração do identificador
    return nome_aba.split('/')[0]
                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                    ZX
def criar_nome_aba(workbook, base_name):
    name = base_name
    count = 1
    while name in workbook.sheetnames:
        count += 1
        name = f"{base_name}{count}"
    return name

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Verifica se o arquivo foi enviado pelo formulário.
        file = request.files['file']

        # Verificar se o usuário escolheu um arquivo ou cancelou a seleção
        if not file:
            return "Nenhum arquivo selecionado. Encerrando o programa."

        # Ler o arquivo CSV com o pandas
        dataframe = pd.read_csv(file)

        # Preencher valores vazios na coluna 'Nome operador (avaliação)'
        dataframe['Nome operador (identificação)'].fillna('Verificar qual operador identificou', inplace=True)
        dataframe['Nome operador (avaliação)'].fillna('Sem operador analisando ainda', inplace=True)

        # Identificar valores únicos na coluna "Macroprocesso"
        macroprocessos = dataframe['Macroprocesso'].unique()

        # Criar um arquivo Excel
        workbook = Workbook()

        # Remover a planilha padrão
        default_sheet = workbook['Sheet']
        workbook.remove(default_sheet)

        abas_criadas = set()  # Conjunto para rastrear abas criadas

        for macroprocesso in macroprocessos:
            # Filtrar o DataFrame para o Macroprocesso atual
            df_macroprocesso = dataframe[dataframe['Macroprocesso'] == macroprocesso]
            unidades_adm = df_macroprocesso['Unidades administrativas análise funcional'].unique()

            for unidade_operador in unidades_adm:
                df_unidade_operador = df_macroprocesso[df_macroprocesso['Unidades administrativas análise funcional'] == unidade_operador]

                # Agrupar os dados por situação e operador de identificação
                agrupado = df_macroprocesso.groupby(
                        ['Processo', 'Subprocesso', 'Situação', 'Nome operador (identificação)', 'Nome operador (avaliação)', 'Unidades administrativas análise funcional']).size().reset_index(
                        name='Qtd. de Docs')

                # Verificar a situação 'ANÁLISE DE AVALIAÇÃO' e criar a coluna 'STATUS' com a mensagem adequada para cada status
                # agrupado['Status'] = ''
                agrupado.loc[agrupado['Situação'] == 'EM IDENTIFICAÇÃO', 'Status'] = 'COM O OPERADOR'
                agrupado.loc[agrupado['Situação'] == 'EM ANÁLISE', 'Status'] = 'ANÁLISE RIVIC'
                agrupado.loc[agrupado['Situação'] == 'APROVADA', 'Status'] = 'PENDENTE'
                agrupado.loc[agrupado['Situação'] == 'EM AVALIAÇÃO', 'Status'] = 'PENDENTE'
                agrupado.loc[agrupado['Situação'] == 'ANÁLISE DE AVALIAÇÃO', 'Status'] = 'ANÁLISE RIVIC'
                agrupado.loc[agrupado['Situação'] == 'APROVAÇÃO DE AVALIAÇÃO', 'Status'] = 'FINALIZADA'
                agrupado.loc[agrupado['Situação'] == 'FINALIZADA', 'Status'] = 'FINALIZADA'

                # Adicionar coluna 'Macroprocesso' com valor único
                agrupado['Macroprocesso'] = macroprocesso

                # Mostrar os usuários atribuídos a cada situação
                for index, row in agrupado.iterrows():
                    macroprocesso = row['Macroprocesso']
                    processo = row['Processo']
                    subproc = row['Subprocesso']
                    operadorid = row['Nome operador (identificação)']
                    unidade = row['Unidades administrativas análise funcional']
                    contagem = row['Qtd. de Docs']
                    situacao = row['Situação']
                    status = row['Status']
                    print(f"Contagem: {contagem}, Situação: {situacao}")

                # Selecionar apenas as colunas desejadas
                colunas_desejadas = ['Macroprocesso', 'Processo', 'Subprocesso', 'Situação', 'Nome operador (identificação)',
                                         'Nome operador (avaliação)', 'Unidades administrativas análise funcional', 'Qtd. de Docs', 'Status']
                agrupado = agrupado[colunas_desejadas]

                aba_nome_base = unidade_operador.split('/')[0][:31]
                sigla = unidade_operador.split('/')[1] if '/' in unidade_operador else ''  # Verificar se há uma barra
                aba_nome = criar_nome_aba(workbook, aba_nome_base, sigla)
                worksheet = criar_aba(workbook, aba_nome, abas_criadas)

                # Pegar parte da Unidade Operador até a primeira barra e limitar a 31 caracteres
                #worksheet = workbook.create_sheet(title=aba_nome)

                # Preencher o arquivo Excel com os dados do DataFrame
                for row in dataframe_to_rows(agrupado, index=False, header=True):
                    worksheet.append(row)

                # Adicionar formatação à coluna 'Qtd. de Docs'
                for row in worksheet.iter_rows(min_row=2, min_col=colunas_desejadas.index('Qtd. de Docs') + 1, max_col=colunas_desejadas.index('Qtd. de Docs') + 1):
                    for cell in row:
                        if cell.value is not None and cell.value <= 4:
                            cell.fill = PatternFill(start_color='FFA07A', end_color='FFA07A',
                                                        fill_type='solid')  # Substitua a cor conforme necessário

                for row in worksheet.iter_rows(min_row=2, min_col=colunas_desejadas.index('Status') + 1, max_col=colunas_desejadas.index('Status') + 1):
                    for cell in row:
                        if cell.value == 'FINALIZADA':
                                cell.fill = PatternFill(start_color='9AFF9A', end_color='9AFF9A', fill_type='solid')
                        if cell.value == 'ANÁLISE RIVIC':
                                cell.fill = PatternFill(start_color='BFEFFF', end_color='9AC0CD', fill_type='solid')
                        if cell.value == 'COM O OPERADOR':
                                cell.fill = PatternFill(start_color='FF7F00', end_color='FF7F00', fill_type='solid')
                        if cell.value == 'PENDENTE':
                                cell.fill = PatternFill(start_color='FF3030', end_color='FF3030', fill_type='solid')

                # Adicionar título para cada coluna no topo da planilha
                for col_num, col_name in enumerate(colunas_desejadas, 1):
                    worksheet.cell(row=1, column=col_num, value=col_name)
                    worksheet.cell(row=1, column=col_num).alignment = Alignment(horizontal='center', vertical='center')

                    # Adicionar título em negrito para cada coluna no topo da planilha
                    for col_num, col_name in enumerate(colunas_desejadas, 1):
                        worksheet.cell(row=1, column=col_num, value=col_name)
                        worksheet.cell(row=1, column=col_num).alignment = Alignment(horizontal='center', vertical='center')
                        worksheet.cell(row=1, column=col_num).font = Font(bold=True)

                # Definir a referência da área com os filtros (todas as colunas, primeira linha até última linha)
                ref = f"A1:{worksheet.cell(row=worksheet.max_row, column=len(colunas_desejadas)).coordinate}"
                worksheet.auto_filter.ref = ref

                # Ajustar a largura das colunas selecionadas com base no conteúdo
                for col in worksheet.iter_cols(min_col=1, max_col=len(colunas_desejadas)):
                    max_length = 0
                    column_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Centralizar o conteúdo das células
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                # Centralizar o conteúdo das células
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # Definir o estilo de borda
                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))

                # Percorrer todas as células da planilha e aplicar o estilo de borda
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = thin_border

                primeira_celula = df_macroprocesso['Macroprocesso'].iloc[0]

        # Salvar o arquivo Excel no servidor externo
        current_datetime = datetime.now().strftime("%Y%m%d")  # Obter a data atual no formato AAAAMMDD
        server_destination = 'C:/Users/Patrick Ribeiro/Documents/dev/Rivic/Pendencias'  # Substitua pelo caminho absoluto no servidor
        filename = f"{current_datetime}_Relatório_Macroprocessos.xlsx"  # Nome do arquivo com prefixo de data
        excel_path = os.path.join(server_destination, filename)
        workbook.save(excel_path)

        # Enviar o arquivo para o cliente como anexo
        return send_file(filename, as_attachment=True)

    return render_template('index.html')  # Crie a página "index.html" no diretório "templates".


@app.route('/gerar_links', methods=['POST'])
def gerar_links():
    if request.method == 'POST':
        # Verifica se o arquivo foi enviado pelo formulário.
        file = request.files['file']

        # Verificar se o usuário escolheu um arquivo ou cancelou a seleção
        if not file:
            return "Nenhum arquivo selecionado. Encerrando o programa."

        # Read the CSV file with pandas
        df = pd.read_csv(file)

        '''# Drop the column (if it exists)
        if "Processo" in df.columns:
            df.drop(columns=["Processo"], inplace=True)
        if "Nome operador (identificação)" in df.columns:
            df.drop(columns=["Nome operador (identificação)"], inplace=True)
        if "Nome operador (avaliação)" in df.columns:
            df.drop(columns=["Nome operador (avaliação)"], inplace=True)
        if "Unidade(s) Operador" in df.columns:
            df.drop(columns=["Unidade(s) Operador"], inplace=True)
        if "AÇÃO" in df.columns:
            df.drop(columns=["AÇÃO"], inplace=True)'''

        # Lista das colunas desejadas
        colunas_desejadas = ["Macroprocesso", "Processo", "Tipo documental", "id"]

        # Selecionar apenas as colunas desejadas no DataFrame
        df = df[colunas_desejadas]

        # Create a new DataFrame for the Excel output
        output_df = df.copy()

        # Create the link column by combining the base link and the ID from the "ID" column
        base_link = "https://app08.virtuaserver.com.br/astrum/public/site/html/tipologiaCadastra.html?&"
        output_df["Link"] = base_link + df["id"].astype(str) + "#"

        # Create a new Excel file using openpyxl
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Document List"

        # Add titles to the first row
        titles = ["MACROPROCESSO", "PROCESSO", "DOCUMENTO", "ID", "LINK"]
        worksheet.append(titles)

        # Write the data from the DataFrame to the Excel file
        for row in dataframe_to_rows(output_df, index=False, header=False):
            worksheet.append(row)

        # Set the header row in bold
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        # Adjust the column width to fit the content
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        # Insert a blank row after the first row
        worksheet.insert_rows(2)

        # Make the "LINK" column clickable
        for row in worksheet.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.hyperlink = cell.value
                cell.font = Font(underline="single", color="0563C1")

        # Save the Excel file on the server
        destination = '/home/rickribeiro/mysite/files/'  # Replace with the absolute path on the server
        filename = 'Links_Download.xlsx'
        excel_path = os.path.join(destination, filename)
        workbook.save(excel_path)

        # Send the file to the client as an attachment
        return send_file(excel_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)