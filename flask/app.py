from flask import Flask, render_template, request, send_file
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Verifica se o arquivo foi enviado pelo formulário.
        file = request.files['file']

        # Verificar se o usuário escolheu um arquivo ou cancelou a seleção
        if not file:
            return "Nenhum arquivo selecionado. Encerrando o programa."

        # Ler o arquivo CSV com o pandas
        dataframe = pd.read_csv(file)

        # Preencher valores vazios na coluna 'Nome operador (avaliação)'
        dataframe['Nome operador (identificação)'].fillna('Verificar qual operador identificou', inplace=True)
        dataframe['Nome operador (avaliação)'].fillna('Sem operador analisando ainda', inplace=True)
        # Agrupar os dados por situação e operador de identificação
        agrupado = dataframe.groupby(['Situação', 'Nome operador (identificação)', 'Nome operador (avaliação)']).size().reset_index(name='Qtd. de Docs')

        # Adicionar coluna 'Status' com valor vazio
        agrupado['Status'] = ''

        # Adicionar coluna 'Macroprocesso' com valor único
        agrupado['Macroprocesso'] = dataframe['Macroprocesso'][0]

        # Mostrar os usuários atribuídos a cada situação
        for index, row in agrupado.iterrows():
            macroprocesso = row['Macroprocesso']
            operadorid = row['Nome operador (identificação)']
            contagem = row['Qtd. de Docs']
            situacao = row['Situação']
            status = row['Status']
            print(f"Contagem: {contagem}, Situação: {situacao}")

        # Selecionar apenas as colunas desejadas
        colunas_desejadas = ['Macroprocesso', 'Situação', 'Nome operador (identificação)', 'Nome operador (avaliação)', 'Qtd. de Docs', 'Status']
        agrupado= agrupado[colunas_desejadas]


        # Verificar a situação 'ANÁLISE DE AVALIAÇÃO' e criar a coluna 'STATUS' com a mensagem adequada para cada status
        #agrupado['Status'] = ''
        agrupado.loc[agrupado['Situação'] == 'ANÁLISE DE AVALIAÇÃO', 'Status'] = 'ANÁLISE RIVIC'
        agrupado.loc[agrupado['Situação'] == 'EM AVALIAÇÃO', 'Status'] = 'PENDENTE'
        agrupado.loc[agrupado['Situação'] == 'APROVADA', 'Status'] = 'PENDENTE'
        agrupado.loc[agrupado['Situação'] == 'APROVAÇÃO DE AVALIAÇÃO', 'Status'] = 'FINALIZADA'

        # Criar um novo arquivo Excel usando o openpyxl
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Relatório de Pendências"

        # Adicionar uma linha adicional para evitar que outras sejam suprimidas
        worksheet.insert_rows(1)

        # Preencher o arquivo Excel com os dados do DataFrame
        for row in agrupado.itertuples(index=False):
            worksheet.append(row)
        for row in worksheet.iter_rows(min_row=2, min_col=colunas_desejadas.index('Status') + 1, max_col=colunas_desejadas.index('Status') + 1):
            for cell in row:
                if cell.value == 'FINALIZADA':
                    cell.fill = PatternFill(start_color='9AFF9A', end_color='9AFF9A', fill_type='solid')
                if cell.value == 'ANÁLISE RIVIC':
                    cell.fill = PatternFill(start_color='FF7F00', end_color='FF7F00', fill_type='solid')
                if cell.value == 'PENDENTE':
                    cell.fill = PatternFill(start_color='FF3030', end_color='FF3030', fill_type='solid')


        # Adicionar título para cada coluna no topo da planilha
        for col_num, col_name in enumerate(colunas_desejadas, 1):
            worksheet.cell(row=1, column=col_num, value=col_name)
            worksheet.cell(row=1, column=col_num).alignment = Alignment(horizontal='center', vertical='center')

        # Adicionar título em negrito para cada coluna no topo da planilha
        for col_num, col_name in enumerate(colunas_desejadas, 1):
            worksheet.cell(row=1, column=col_num, value=col_name)
            worksheet.cell(row=1, column=col_num).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=1, column=col_num).font = Font(bold=True)

        # Definir a referência da área com os filtros (todas as colunas, primeira linha até última linha)
        ref = f"A1:{worksheet.cell(row=worksheet.max_row, column=len(colunas_desejadas)).coordinate}"
        worksheet.auto_filter.ref = ref

        # Ajustar a largura das colunas selecionadas com base no conteúdo
        for col in worksheet.iter_cols(min_col=1, max_col=len(colunas_desejadas)):
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Centralizar o conteúdo das células
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        primeira_celula = dataframe['Macroprocesso'][0]

        # Definir o estilo de borda
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Percorrer todas as células da planilha e aplicar o estilo de borda
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border

        # Criar uma nova aba para os operadores e suas respectivas gerências
        aba_operadores = workbook.create_sheet(title="Operadores e Gerências")

        # Preencher a nova aba com os dados de operadores e suas respectivas gerências
        operadores_gerencias = dataframe[['Nome operador (identificação)', 'Nome operador (avaliação)', 'Unidade(s) Operador']]

        #adicionar uma nova linha ao cabeçalho
        aba_operadores.append(["Nome Operador (Identificação)", "Nome Operador (Avaliação)", "Gerência"])
        aba_operadores.append([])

        for row in operadores_gerencias.itertuples(index=False):
            aba_operadores.append(row)

        # Ajustar a largura das colunas na nova aba
        for col_num, col_name in enumerate(operadores_gerencias.columns, 1):
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in aba_operadores[get_column_letter(col_num)])
            adjusted_width = (max_length + 2) * 1.2
            aba_operadores.column_dimensions[get_column_letter(col_num)].width = adjusted_width

        # Centralizar o conteúdo das células na nova aba
        for row in aba_operadores.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # Adicionar título em negrito para cada coluna no topo da planilha
        for col_num, col_name in enumerate(operadores_gerencias.columns, 1):
            aba_operadores.cell(row=1, column=col_num, value=col_name)
            aba_operadores.cell(row=1, column=col_num).alignment = Alignment(horizontal='left', vertical='center')
            aba_operadores.cell(row=1, column=col_num).font = Font(bold=True)

        # Definir a referência da área com os filtros (todas as colunas, primeira linha até última linha)
        ref = f"A1:{aba_operadores.cell(row=aba_operadores.max_row, column=len(operadores_gerencias)).coordinate}"
        aba_operadores.auto_filter.ref = ref

        # Aplicar o mesmo código de formatação das células à nova aba "Operadores e Gerências"
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for row in aba_operadores.iter_rows():
            for cell in row:
                cell.border = thin_border

        # Salvar o arquivo Excel
        workbook.save(f'{primeira_celula}_PENDENCIAS.xlsx')
        return send_file(f'{primeira_celula}_PENDENCIAS.xlsx', as_attachment=True)
    return render_template('index.html')  # Crie a página "index.html" no diretório "templates".
if __name__ == '__main__':
    app.run(debug=True)